from openpyxl import load_workbook, Workbook
from phonenumbers import parse, format_number, PhoneNumberFormat


# прочитали из файла
wb = load_workbook("usersdescM.xlsx")
ws = wb.get_sheet_by_name("usersdescM")
col = ws["C"]

# получили все телефоны
phones_list = []
for i in col:
    phones_list.append(i.value)


def normalize_number(num: str):
    number = parse(num, "RU")
    return format_number(number, PhoneNumberFormat.E164)


# нормализовали значения
norm_list = []

for num in phones_list:
    if num != None:
        norm_list.append(normalize_number(str(num)))
    if num == None:
        norm_list.append(num)

valid_num = 0
invalid_num = 0
zero = 0
for num in norm_list:
    if num == None:
        zero += 1
        continue
    if len(num) == 12:
        valid_num += 1
    if len(num) != 12:
        invalid_num += 1
    if len(num) > 12:
        print(num)
# print(f'валидных номеров {valid_num}, не валидных {invalid_num}, нулов {zero}, всего {valid_num + invalid_num + zero}')
# print(f'исходный список {len(phones_list)}, нормализованый список {len(norm_list)}')


result = []
counter = 0
# привели к одному виду
for num in norm_list:
    if num == None:
        result.append(num)
        continue
    if len(num) == 12:
        valid = (
            num[0:2]
            + "("
            + num[2:5]
            + ")"
            + num[5:8]
            + "-"
            + num[8:10]
            + "-"
            + num[10:12]
        )
        result.append(valid)
        counter += 1
    if len(num) != 12:
        num = None
        result.append(num)

wb1 = Workbook()
ws1 = wb1.active


for i, num in enumerate(result):
    ws1[f"A{i+1}"] = num


wb1.save("phones.xlsx")
